import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Avg, Max, Min, Count, Q

from exams.models import Exam, ExamRecord, AnswerDetail


@login_required
def dashboard_home(request):
    if not request.user.is_admin():
        messages.error(request, '权限不足')
        return redirect('exams:exam_list')

    exams = Exam.objects.annotate(
        record_count=Count('records', filter=Q(records__is_submitted=True)),
        avg_score=Avg('records__total_score', filter=Q(records__is_submitted=True)),
    ).order_by('-created_at')

    return render(request, 'dashboard/home.html', {'exams': exams})


@login_required
def exam_stats(request, exam_id):
    if not request.user.is_admin():
        messages.error(request, '权限不足')
        return redirect('exams:exam_list')

    exam = get_object_or_404(Exam, id=exam_id)
    records = ExamRecord.objects.filter(exam=exam, is_submitted=True).select_related('user')

    stats = records.aggregate(
        avg_score=Avg('total_score'),
        max_score=Max('total_score'),
        min_score=Min('total_score'),
        total_count=Count('id'),
    )

    pass_count = records.filter(total_score__gte=exam.pass_score).count()
    stats['pass_rate'] = round(pass_count / stats['total_count'] * 100, 1) if stats['total_count'] else 0
    stats['pass_count'] = pass_count
    stats['fail_count'] = stats['total_count'] - pass_count

    score_ranges = [
        ('90-100', 90, 101),
        ('80-89', 80, 90),
        ('70-79', 70, 80),
        ('60-69', 60, 70),
        ('0-59', 0, 60),
    ]
    score_distribution = []
    for label, low, high in score_ranges:
        cnt = records.filter(total_score__gte=low, total_score__lt=high).count()
        score_distribution.append({'label': label, 'count': cnt})

    question_stats = []
    for q in exam.questions.all():
        total_answers = AnswerDetail.objects.filter(
            record__exam=exam, record__is_submitted=True, question=q
        ).count()
        correct_answers = AnswerDetail.objects.filter(
            record__exam=exam, record__is_submitted=True, question=q, is_correct=True
        ).count()
        rate = round(correct_answers / total_answers * 100, 1) if total_answers else 0
        question_stats.append({
            'question': q,
            'correct_rate': rate,
            'total': total_answers,
            'correct': correct_answers,
        })

    return render(request, 'dashboard/exam_stats.html', {
        'exam': exam,
        'records': records,
        'stats': stats,
        'score_distribution': score_distribution,
        'question_stats': question_stats,
    })


@login_required
def export_scores(request, exam_id):
    if not request.user.is_admin():
        messages.error(request, '权限不足')
        return redirect('exams:exam_list')

    exam = get_object_or_404(Exam, id=exam_id)
    records = ExamRecord.objects.filter(
        exam=exam, is_submitted=True
    ).select_related('user').order_by('-total_score')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '成绩汇总'

    headers = ['排名', '用户名', '姓名', '供应商', '得分', '是否及格', '开始时间', '交卷时间']
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for rank, record in enumerate(records, 1):
        ws.append([
            rank,
            record.user.username,
            record.user.first_name or record.user.username,
            record.user.department,
            record.total_score,
            '及格' if record.total_score >= exam.pass_score else '不及格',
            record.started_at.strftime('%Y-%m-%d %H:%M:%S') if record.started_at else '',
            record.submitted_at.strftime('%Y-%m-%d %H:%M:%S') if record.submitted_at else '',
        ])

    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 30)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'{exam.title}_成绩汇总.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
